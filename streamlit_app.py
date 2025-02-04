import streamlit as st
import gspread
from datetime import datetime, timedelta
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import random
import pandas as pd


# Configuración inicial de la página (debe ser la primera llamada)
st.set_page_config(
    page_title="Gestor de Playlists 24h",
    page_icon="🎧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --------------------------
# 1. Configuración del Tema
# --------------------------
def setup_theme():
    if 'theme' not in st.session_state:
        st.session_state.theme = "light"  # Tema predeterminado

    # Cambiar el tema según la selección del usuario
    with st.sidebar:
        st.markdown("---")
        st.subheader("🎨 Configuración del Tema")
        theme = st.radio(
            "Selecciona el tema:",
            ["Claro ☀️", "Oscuro 🌙"],
            index=0 if st.session_state.theme == "light" else 1
        )
        if theme == "Oscuro 🌙":
            st.session_state.theme = "dark"
        else:
            st.session_state.theme = "light"


# --------------------------
# 1. Sistema de Login Simple
# --------------------------
def check_login():
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
        
    if not st.session_state.logged_in:
        with st.container():
            col1, col2, col3 = st.columns([1,3,1])
            with col2:
                st.title("🔐 Acceso al Sistema")
                user = st.text_input("Usuario")
                password = st.text_input("Contraseña", type="password")
                
                if st.button("Ingresar", use_container_width=True):
                    if user == "admin" and password == "admin123":
                        st.session_state.logged_in = True
                        st.rerun()  # Usar st.rerun() en lugar de st.experimental_rerun()
                    else:
                        st.error("Credenciales incorrectas")
        st.stop()  # Detener la ejecución si no se ha iniciado sesión


# Función para autenticar Google Sheets usando Streamlit Secrets
def authenticate_google_sheets():
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        
        # Obtener credenciales desde st.secrets
        credentials_dict = st.secrets["google_sheets"]
        
        # Crear credenciales desde el diccionario
        credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
        client = gspread.authorize(credentials)
        
        return client
    except Exception as e:
        st.error(f"Error al autenticar Google Sheets: {e}")
        return None


# Función para cargar programas desde Google Sheets
def load_programs_from_google_sheet():
    client = authenticate_google_sheets()
    if not client:
        return []
    try:
        spreadsheet_id = '1Ka9YhP860lZlibXudUkr7an7zGs-spO54KBmidpNr1A'
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.sheet1
        data = worksheet.get_all_records()
        programs = [{'name': row['Name'], 'duration': row['Duration']} for row in data]
        st.session_state.messages.append({"type": "success", "content": "Programas cargados correctamente"})
        return programs
    except Exception as e:
        st.session_state.messages.append({"type": "error", "content": f"Error al cargar programas: {e}"})
        return []

# Función para cargar promos desde Google Sheets
def load_promos_from_google_sheet():
    client = authenticate_google_sheets()
    if not client:
        return []
    try:
        spreadsheet_id = '17AtkM82WEWczbzLvHSq-XYQbiAImTNkmSguDlDg_46g'
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.sheet1
        data = worksheet.get_all_records()
        promos = []
        for row in data:
            try:
                h, m, s = map(int, row['Duration'].split(':'))
                duration_seconds = h * 3600 + m * 60 + s
                promos.append({'name': row['Name'], 'duration': duration_seconds})
            except ValueError:
                st.session_state.messages.append({"type": "error", "content": f"Error al procesar la duración de la promo '{row['Name']}'. Formato inválido: {row['Duration']}"})
        st.session_state.messages.append({"type": "success", "content": "Promos cargadas correctamente"})
        return promos
    except Exception as e:
        st.session_state.messages.append({"type": "error", "content": f"Error al cargar promos: {e}"})
        return []

# Función para cargar rellenos desde una hoja específica de Google Sheets
def load_fillers_from_google_sheet(sheet_name):
    client = authenticate_google_sheets()
    if not client:
        return []
    try:
        spreadsheet_id = '1MjcPISQEPUvYAHqVtW7nvweqfXhaS_cAbREjeG3uK-I'
        spreadsheet = client.open_by_key(spreadsheet_id)
        worksheet = spreadsheet.worksheet(sheet_name)  # Seleccionar la hoja por nombre
        data = worksheet.get_all_records()
        fillers = []
        for row in data:
            try:
                h, m, s = map(int, row['Duration'].split(':'))
                duration_seconds = h * 3600 + m * 60 + s
                fillers.append({'name': row['Name'], 'duration': duration_seconds})
            except ValueError:
                st.session_state.messages.append({"type": "error", "content": f"Error al procesar la duración del relleno '{row['Name']}'. Formato inválido: {row['Duration']}"})
        st.session_state.messages.append({"type": "success", "content": f"Rellenos cargados correctamente desde la hoja: {sheet_name}"})
        return fillers
    except Exception as e:
        st.session_state.messages.append({"type": "error", "content": f"Error al cargar rellenos: {e}"})
        return []

# Función para listar las hojas disponibles en el Google Sheet
def list_sheets():
    client = authenticate_google_sheets()
    if not client:
        return []
    try:
        spreadsheet_id = '1MjcPISQEPUvYAHqVtW7nvweqfXhaS_cAbREjeG3uK-I'
        spreadsheet = client.open_by_key(spreadsheet_id)
        sheets = [sheet.title for sheet in spreadsheet.worksheets()]
        return sheets
    except Exception as e:
        st.session_state.messages.append({"type": "error", "content": f"Error al listar las hojas: {e}"})
        return []

# Función para exportar a Excel
def export_to_excel(playlist):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Playlist"
        headers = ['Item', 'Hora de Inicio', 'Nombre', 'Duración', 'Tipo']
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill
        for i, block in enumerate(playlist, start=1):
            ws.append([block['item'], block['start_time'], block['name'], block['duration'], block['type']])
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2
        filename = f"playlist_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        wb.save(filename)
        st.session_state.messages.append({"type": "success", "content": f"Playlist exportada correctamente a: {filename}"})
    except Exception as e:
        st.session_state.messages.append({"type": "error", "content": f"Error al exportar a Excel: {e}"})

# Función para exportar a Google Sheets con colores
def export_to_google_sheets(playlist, sheet_title):
    try:
        client = authenticate_google_sheets()
        if not client:
            return
        spreadsheet_id = '1SeKSZLR7IWrVVj9ny5hezcS-Nro06Amp9S29W6pMovU'  # Reemplaza con tu ID
        spreadsheet = client.open_by_key(spreadsheet_id)

        # Crear una nueva hoja dentro del Google Sheet
        worksheet = spreadsheet.add_worksheet(title=sheet_title, rows="100", cols="5")

        # Escribir los encabezados en la primera fila
        headers = ['Item', 'Hora de Inicio', 'Nombre', 'Duración', 'Tipo']
        worksheet.update(values=[headers], range_name='A1:E1')

        # Definir colores para cada tipo
        type_colors = {
            'Program': {'red': 0.8, 'green': 0.8, 'blue': 0.2},  # Amarillo
            'Tanda': {'red': 0.2, 'green': 0.8, 'blue': 0.2},    # Verde
            'Promo': {'red': 0.9, 'green': 0.6, 'blue': 0.1},    # Naranja
            'Filler': {'red': 0.5, 'green': 0.5, 'blue': 0.5},   # Gris
        }

        # Crear las filas de datos
        rows = []
        formats = []
        for i, block in enumerate(playlist, start=2):  # Comenzar desde la fila 2
            rows.append([
                block['item'],  # Número de ítem
                block['start_time'],
                block['name'],
                block['duration'],
                block['type']
            ])
            # Aplicar formato de color según el tipo
            row_range = f'A{i}:E{i}'
            formats.append({
                "range": row_range,
                "format": {
                    "backgroundColor": type_colors.get(block['type'], {'red': 1, 'green': 1, 'blue': 1}),
                    "textFormat": {"bold": block['type'] in ['Program', 'Tanda']}
                }
            })

        # Escribir las filas en una sola llamada
        worksheet.update(values=rows, range_name=f'A2:E{len(rows) + 1}')

        # Aplicar los formatos en un solo lote
        worksheet.batch_format(formats)

        # Aplicar formato a los encabezados
        worksheet.format('A1:E1', {
            'backgroundColor': {'red': 0.0, 'green': 0.5, 'blue': 0.8},  # Azul claro
            'textFormat': {'bold': True, 'foregroundColor': {'red': 1.0, 'green': 1.0, 'blue': 1.0}}  # Blanco
        })

        st.session_state.messages.append({"type": "success", "content": f"Playlist exportada correctamente a Google Sheets: {spreadsheet.url} -> {sheet_title}"})
    except Exception as e:
        st.session_state.messages.append({"type": "error", "content": f"Error al exportar a Google Sheets: {e}"})

# Función para generar la playlist
def generate_playlist(start_time, end_time, promos, fillers, user_programs):
    current_time = start_time
    playlist = []
    block_counter = 1
    user_program_index = 0
    item_counter = 1

    # Calcular el tiempo total de la playlist
    total_time = (end_time - start_time).total_seconds()
    elapsed_time = 0

    # Barra de progreso
    progress_bar = st.progress(0)
    status_text = st.empty()  # Para mostrar el estado actual

    # Agregar tanda de 60 segundos al inicio
    tanda_duration = 60
    playlist.append({
        "item": item_counter,
        "start_time": current_time.strftime("%H:%M:%S"),
        "name": "Tanda 60 segundos",
        "duration": str(timedelta(seconds=tanda_duration)),
        "type": "Tanda",
        "block": block_counter
    })
    item_counter += 1
    current_time += timedelta(seconds=tanda_duration)
    elapsed_time += tanda_duration

    # Actualizar barra de progreso
    progress = min(elapsed_time / total_time, 1.0)
    progress_bar.progress(progress)
    status_text.text(f"Generando playlist... {int(progress * 100)}% completado")

    while current_time < end_time:
        if user_program_index < len(user_programs):
            program = user_programs[user_program_index]
            program_duration = parse_duration(program["duration"])
            playlist.append({
                "item": item_counter,
                "start_time": current_time.strftime("%H:%M:%S"),
                "name": program["name"],
                "duration": program["duration"],
                "type": "Program",
                "block": block_counter
            })
            item_counter += 1
            current_time += timedelta(seconds=program_duration)
            elapsed_time += program_duration
            user_program_index += 1

            # Actualizar barra de progreso
            progress = min(elapsed_time / total_time, 1.0)
            progress_bar.progress(progress)
            status_text.text(f"Generando playlist... {int(progress * 100)}% completado")
        else:
            break

        # Agregar tanda de 60 segundos después del programa
        tanda_duration = 60
        playlist.append({
            "item": item_counter,
            "start_time": current_time.strftime("%H:%M:%S"),
            "name": "Tanda 60 segundos",
            "duration": str(timedelta(seconds=tanda_duration)),
            "type": "Tanda",
            "block": block_counter
        })
        item_counter += 1
        current_time += timedelta(seconds=tanda_duration)
        elapsed_time += tanda_duration

        # Actualizar barra de progreso
        progress = min(elapsed_time / total_time, 1.0)
        progress_bar.progress(progress)
        status_text.text(f"Generando playlist... {int(progress * 100)}% completado")

        # Calcular tiempo restante en el bloque
        time_to_next_block = calculate_time_to_next_block(current_time)
        remaining_time = time_to_next_block.seconds

        # Agregar promos y rellenos
        if remaining_time > 0:
            selected_content = select_content(remaining_time, promos + fillers)
            for content in selected_content:
                playlist.append({
                    "item": item_counter,
                    "start_time": current_time.strftime("%H:%M:%S"),
                    "name": content['name'],
                    "duration": str(timedelta(seconds=content['duration'])),
                    "type": "Promo" if content in promos else "Filler",
                    "block": block_counter
                })
                item_counter += 1
                current_time += timedelta(seconds=content['duration'])
                elapsed_time += content['duration']

                # Actualizar barra de progreso
                progress = min(elapsed_time / total_time, 1.0)
                progress_bar.progress(progress)
                status_text.text(f"Generando playlist... {int(progress * 100)}% completado")

            if remaining_time > 0:
                playlist.append({
                    "item": item_counter,
                    "start_time": current_time.strftime("%H:%M:%S"),
                    "name": "Tanda Parcial",
                    "duration": str(timedelta(seconds=remaining_time)),
                    "type": "Tanda",
                    "block": block_counter
                })
                item_counter += 1
                current_time += timedelta(seconds=remaining_time)
                elapsed_time += remaining_time

                # Actualizar barra de progreso
                progress = min(elapsed_time / total_time, 1.0)
                progress_bar.progress(progress)
                status_text.text(f"Generando playlist... {int(progress * 100)}% completado")

        block_counter += 1

    # Finalizar barra de progreso
    progress_bar.progress(1.0)
    status_text.text("Playlist generada exitosamente 🎉")

    return playlist

# Función para convertir duración en formato HH:MM:SS a segundos
def parse_duration(duration_str):
    h, m, s = map(int, duration_str.split(':'))
    return h * 3600 + m * 60 + s

# Función para calcular el tiempo hasta el siguiente bloque
def calculate_time_to_next_block(current_time):
    valid_start_minutes = [0, 10, 15, 20, 30, 40, 45, 50, 0]
    current_minute = current_time.minute
    next_minute = next((m for m in valid_start_minutes if m > current_minute), valid_start_minutes[0])
    next_block_time = current_time.replace(minute=next_minute, second=0, microsecond=0)
    if next_block_time <= current_time:
        next_block_time += timedelta(hours=1)
    return next_block_time - current_time

# Función para seleccionar contenido
def select_content(duration_seconds, content_list):
    random.shuffle(content_list)
    selected = []
    remaining_seconds = duration_seconds
    content_list.sort(key=lambda x: x['duration'], reverse=True)
    for content in content_list:
        if content['duration'] <= remaining_seconds:
            selected.append(content)
            remaining_seconds -= content['duration']
        if remaining_seconds <= 0:
            break
    return selected

# Interfaz de Streamlit
def main():
    # Configurar el tema
    setup_theme()

    # Verificar el login
    check_login()
    
    # Inicializar estados
    if 'playlist' not in st.session_state:
        st.session_state.playlist = None
    if 'sheet_title' not in st.session_state:
        st.session_state.sheet_title = f"Playlist_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"
    if 'messages' not in st.session_state:
        st.session_state.messages = []
    if 'programs' not in st.session_state:
        st.session_state.programs = []

    # ------------------------------------------------------
    # Barra Lateral (Todo el contenido del sidebar aquí)
    # ------------------------------------------------------
    with st.sidebar:
        # Sección de Mensajes Importantes (Parte superior)
        st.header("📢 Notificaciones")
        messages_container = st.container(height=200)
        with messages_container:
            for msg in st.session_state.messages[-3:]:  # Mostrar últimos 3 mensajes
                if msg["type"] == "success":
                    st.success(msg["content"], icon="✅")
                elif msg["type"] == "error":
                    st.error(msg["content"], icon="❌")
                elif msg["type"] == "warning":
                    st.warning(msg["content"], icon="⚠️")

        st.markdown("---")
        
        # Sección Principal de Configuración
        st.header("⚙️ Configuración Principal")
        
        # Selector de hoja de rellenos
        sheets = list_sheets()
        selected_sheet = st.selectbox(
            "📂 Seleccionar hoja de rellenos:", 
            sheets if sheets else ["No disponible"],
            disabled=not sheets
        )
        
        st.markdown("---")
        
        # Sección de Horarios (Configuración menos importante)
        st.subheader("⏰ Configuración de Horarios")
        start_time = st.time_input("Hora de inicio", value=datetime.strptime("05:59:00", "%H:%M:%S").time())
        end_time = st.time_input("Hora de fin", value=datetime.strptime("23:59:00", "%H:%M:%S").time())

        # Mostrar la lista de programas en una tabla
        st.markdown("---")
        st.subheader("📋 Lista de Programas")
        if st.session_state.programs:
            # Convertir la lista de programas en un DataFrame
            programs_df = pd.DataFrame(st.session_state.programs)
            # Mostrar la tabla en el sidebar
            st.dataframe(programs_df, use_container_width=True, hide_index=True)
        else:
            st.write("No hay programas cargados.")

    # ------------------------------------------------------
    # Cuerpo Principal de la Aplicación
    # ------------------------------------------------------
    st.title("🎵 Generador de Playlist")
    st.markdown("---")

    # Cargar datos
    with st.spinner("🔍 Cargando programas y promos..."): 
        promos = load_promos_from_google_sheet()
        user_programs = load_programs_from_google_sheet()
        fillers = load_fillers_from_google_sheet(selected_sheet) if sheets else []

        # Actualizar la lista de programas en el estado de la sesión
        st.session_state.programs = user_programs

    # Generar playlist
    col1, col2 = st.columns([1,3])
    with col1:
        if st.button("🎶 Generar Playlist", type="primary", help="Genera una nueva playlist basada en los parámetros actuales"):
            if not user_programs or not promos or not fillers:
                st.session_state.messages.append({"type": "warning", "content": "Faltan datos para generar la playlist"})
            else:
                start_time_dt = datetime.combine(datetime.today(), start_time)
                end_time_dt = datetime.combine(datetime.today(), end_time)
                playlist = generate_playlist(start_time_dt, end_time_dt, promos, fillers, user_programs)
                st.session_state.playlist = playlist
                st.session_state.messages.append({"type": "success", "content": "Playlist generada correctamente"})

    # Vista previa de playlist
    if st.session_state.playlist:
        st.markdown("### 📜 Vista Previa")
        st.dataframe(
            st.session_state.playlist,
            column_config={
                "item": "Ítem",
                "start_time": {"label": "Hora Inicio", "help": "Hora de inicio del bloque"},
                "name": "Contenido",
                "duration": "Duración",
                "type": {"label": "Tipo", "help": "Tipo de contenido (Programa, Tanda, etc.)"}
            },
            use_container_width=True,
            hide_index=True
        )
        
        # Sección de Exportación
        st.markdown("---")
        st.markdown("### 📤 Exportar Playlist")
        
        col_export1, col_export2 = st.columns(2)
        with col_export1:
            new_sheet_name = st.text_input(
                "📝 Nombre para la hoja:",
                value=st.session_state.sheet_title,
                help="Nombre que tendrá la hoja en Google Sheets"
            )
            st.session_state.sheet_title = new_sheet_name
            
        with col_export2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)  # Espaciado
            if st.button("💾 Exportar a Google Sheets", use_container_width=True):
                if st.session_state.playlist:
                    export_to_google_sheets(st.session_state.playlist, st.session_state.sheet_title)
                else:
                    st.session_state.messages.append({"type": "error", "content": "No hay playlist para exportar"})
            
            if st.button("📥 Exportar a Excel", use_container_width=True):
                if st.session_state.playlist:
                    export_to_excel(st.session_state.playlist)
                else:
                    st.session_state.messages.append({"type": "error", "content": "No hay playlist para exportar"})

if __name__ == "__main__":
    main()